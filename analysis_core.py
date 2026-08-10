"""
Analysis core for the Mailbox Knowledge Gap Analyzer.

Everything in this module is UI-free: no Streamlit imports, no session
state. That separation exists so the logic that produces client-facing
numbers can be unit-tested (see tests/test_core.py) and reused outside
the Streamlit shell. app.py owns flow and presentation only.
"""
import functools
import io
import re

import pandas as pd

from llm_classifier import (
    classify_batch_with_llm,
    make_client,
    BATCH_SIZE,
)

# ============================================================
# Constants
# ============================================================
DEFAULT_TOPICS = {
    "Password Reset": r"password|locked out|reset",
    "VPN / Remote Access": r"\bvpn\b|remote access|certificate",
    "Software Install / Licensing": r"licen[cs]e|install|software",
    "Printer / Hardware": r"printer|mouse|keyboard|laptop|monitor|hardware",
    "New Starter / Access Provisioning": r"new starter|new hire|access request|offboard|provisioning",
    "Email / Outlook Issues": r"outlook|mailbox|calendar|junk folder|spam",
    "WiFi / Network": r"wi-?fi|network|internet connection",
    "Meeting Room / AV": r"conference room|boardroom|\bav\b|screen",
    "Security / Phishing": r"phishing|suspicious email|security policy",
}

QUOTE_MARKERS = re.compile(
    r"(From:\s|Sent:\s|-{3,}\s*Original Message\s*-{3,}|On .{5,80}wrote:|"
    r"________________________________)", re.I)
QUESTION_PATTERN = re.compile(
    r"\?|how do i|how does|what is|what does|where do i|where can i|when is|when do i|"
    r"who is|who do i|can you|could you|please advise|please confirm|not sure|unsure|"
    r"clarif|does this mean|do i need|am i required|is it possible|what happens if|"
    r"could i|can i", re.I)
ACK_PATTERN = re.compile(
    r"^(thanks|thank you|noted|no problem|great|perfect|all good|done|sounds good)\b", re.I)

# NOTE: "service desk" was deliberately removed from this list. In a shared
# mailbox export that includes the team's own outbound replies, those replies
# are written by real people; marking them "automated" skewed the mailbox
# composition statistics. The remaining markers indicate genuine machine
# senders.
AUTOMATED_SENDER_MARKERS = ["no-reply", "noreply", "system", "automated", "notification"]

NON_PERSON_TERMS = {
    "wifi", "wi-fi", "vpn", "outlook", "windows", "microsoft", "teams", "zoom",
    "printer", "password", "mailbox", "internet", "software", "hardware", "helpdesk",
    "service desk", "sharepoint", "onedrive", "excel", "word", "powerpoint",
}


# ============================================================
# File ingestion
# ============================================================
def parse_upload(file_bytes, filename, read_limit=None):
    """Parse an uploaded mailbox export into a DataFrame.

    CSV: pandas nrows genuinely caps what is read from disk.
    XLSX: pandas' default reader materialises every cell of the workbook
    even when nrows is set, which is what allowed a large workbook to
    exhaust hosting memory. Rows are instead streamed through openpyxl in
    read-only mode and reading stops at the cap. (The workbook's shared
    strings table is still parsed by openpyxl -- a pathological file can
    still be heavy, which is why the interface recommends CSV for very
    large mailboxes.)
    XLS: legacy format, handled by xlrd via pandas; these files are
    small by construction (65k-row format limit).
    """
    name = filename.lower()
    buffer = io.BytesIO(file_bytes)
    if name.endswith(".csv"):
        return pd.read_csv(buffer, nrows=read_limit)
    if name.endswith(".xls"):
        return pd.read_excel(buffer, nrows=read_limit, engine="xlrd")

    from openpyxl import load_workbook
    workbook = load_workbook(buffer, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return pd.DataFrame()
        # Build stable, unique column names even when header cells are
        # blank or duplicated -- duplicate names break column detection.
        columns, seen = [], {}
        for i, cell in enumerate(header):
            label = str(cell).strip() if cell is not None and str(cell).strip() else f"Column {i + 1}"
            if label in seen:
                seen[label] += 1
                label = f"{label} ({seen[label]})"
            else:
                seen[label] = 1
            columns.append(label)
        data = []
        for row in rows:
            data.append(row[:len(columns)])
            if read_limit and len(data) >= read_limit:
                break
        return pd.DataFrame(data, columns=columns)
    finally:
        workbook.close()


# ============================================================
# Text preparation
# ============================================================
def strip_html(text):
    """Removes HTML tags and common entities -- many real mailbox exports
    store the body as HTML, and leaving tags in place corrupts every
    downstream step (quote detection, question detection, topic matching)."""
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"&nbsp;|&amp;|&lt;|&gt;|&#\d+;", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def strip_quotes(body_text):
    """Removes quoted reply history, keeping only the new message.
    Falls back safely if the quote appears BEFORE the new text instead of
    after it (common outside Outlook's top-posting convention) -- without
    this check, that ordering would silently delete the real message."""
    match = QUOTE_MARKERS.search(body_text)
    if not match:
        return body_text
    before = body_text[:match.start()].strip()
    after = body_text[match.end():].strip()
    if len(before) >= 10:
        return before
    if len(after) >= 10:
        return after
    return body_text


def classify_email_type(new_content, sender_name):
    text = new_content.strip()
    sender_lc = (sender_name or "").lower()
    if any(m in sender_lc for m in AUTOMATED_SENDER_MARKERS):
        return "automated"
    if not text:
        return "unclear"
    if ACK_PATTERN.match(text) and len(text) < 120:
        return "acknowledgment"
    if QUESTION_PATTERN.search(text):
        return "genuine_question"
    return "status_update"


def tag_topics(text, topic_patterns):
    text_lc = text.lower()
    return [name for name, pattern in topic_patterns.items() if re.search(pattern, text_lc, re.I)]


# ============================================================
# PII redaction
# ============================================================
@functools.lru_cache(maxsize=1)
def get_presidio_analyzer():
    from presidio_analyzer import AnalyzerEngine
    from presidio_analyzer.nlp_engine import NlpEngineProvider
    configuration = {
        "nlp_engine_name": "spacy",
        "models": [{"lang_code": "en", "model_name": "en_core_web_sm"}],
    }
    nlp_engine = NlpEngineProvider(nlp_configuration=configuration).create_engine()
    return AnalyzerEngine(nlp_engine=nlp_engine, supported_languages=["en"])


def _drop_overlapping(results):
    """Presidio can return overlapping spans (e.g. a PERSON overlapping an
    EMAIL_ADDRESS). Replacing overlapping spans corrupts the text into
    hybrid placeholder fragments. Keep the highest-scoring span in any
    overlapping group and drop the rest."""
    accepted = []
    for result in sorted(results, key=lambda r: (-r.score, r.start)):
        if all(result.end <= a.start or result.start >= a.end for a in accepted):
            accepted.append(result)
    return accepted


def redact_dataframe(df, text_columns, progress_callback=None):
    try:
        analyzer = get_presidio_analyzer()
    except Exception:
        analyzer = None
    pseudonym_map = {}
    counters = {}

    def get_pseudonym(original_text, entity_type):
        key = (entity_type, original_text.strip().lower())
        if key not in pseudonym_map:
            counters[entity_type] = counters.get(entity_type, 0) + 1
            pseudonym_map[key] = f"[{entity_type}_{counters[entity_type]}]"
        return pseudonym_map[key]

    def redact_text(text):
        if not text:
            return text
        if analyzer is None:
            # Deployment fallback when Presidio's NLP model is unavailable:
            # emails and phone numbers only. IMPORTANT: this mode does NOT
            # protect personal names -- the caller receives that fact via
            # the third return value and must disclose it to the user.
            matches = []
            patterns = {
                "EMAIL_ADDRESS": r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b",
                "PHONE_NUMBER": r"(?<!\w)(?:\+?\d[\d ()-]{7,}\d)(?!\w)",
            }
            for entity_type, pattern in patterns.items():
                for match in re.finditer(pattern, text, re.I):
                    matches.append((match.start(), match.end(), entity_type))
            results = sorted(matches, key=lambda r: r[0], reverse=True)
            out = text
            for start, end, entity_type in results:
                out = out[:start] + get_pseudonym(text[start:end], entity_type) + out[end:]
            return out
        results = analyzer.analyze(text=text, entities=["PERSON", "EMAIL_ADDRESS", "PHONE_NUMBER"], language="en")
        # Named-entity models can confidently mistake product or technology
        # names (for example "WiFi") for people. Keep email/phone recognizers,
        # but apply a quality threshold and domain guardrail to PERSON results.
        filtered_results = []
        for result in results:
            candidate = text[result.start:result.end].strip()
            if result.entity_type == "PERSON":
                if result.score < 0.70:
                    continue
                if candidate.lower() in NON_PERSON_TERMS:
                    continue
                if not re.fullmatch(r"[A-Za-z][A-Za-z .'-]{1,79}", candidate):
                    continue
                if candidate.isupper() and len(candidate) <= 6:
                    continue
            filtered_results.append(result)
        results = sorted(_drop_overlapping(filtered_results), key=lambda r: r.start, reverse=True)
        out = text
        for r in results:
            placeholder = get_pseudonym(text[r.start:r.end], r.entity_type)
            out = out[:r.start] + placeholder + out[r.end:]
        return out

    total = len(df) * len(text_columns)
    done = 0
    for col in text_columns:
        redacted_col = []
        for val in df[col]:
            redacted_col.append(redact_text(str(val)))
            done += 1
            if progress_callback and done % 5 == 0:
                progress_callback(done / total)
        df[col + "_redacted"] = redacted_col
    if progress_callback:
        progress_callback(1.0)
    # Third return value: True only when the full name-detection engine ran.
    # When False, only email addresses and phone numbers were protected and
    # the interface MUST say so -- silently claiming full protection when
    # names may remain in the text would be a false privacy statement.
    return df, pseudonym_map, analyzer is not None


# ============================================================
# Rule-based analysis (free path)
# ============================================================
def run_analysis(df, topic_patterns, subject_col, body_col, sender_col):
    df = df.copy()
    df["new_content"] = df[body_col].fillna("").astype(str).apply(strip_html).apply(strip_quotes)
    if not sender_col or sender_col not in df.columns:
        sender_values = pd.Series("", index=df.index)
    else:
        sender_values = df[sender_col].fillna("").astype(str)
    df["email_type"] = [classify_email_type(content, sender) for content, sender in zip(df["new_content"], sender_values)]
    df["topics"] = (df[subject_col].fillna("").astype(str) + " " + df["new_content"]).apply(
        lambda t: tag_topics(t, topic_patterns))

    total = len(df)
    rows = []
    for topic in topic_patterns:
        topic_mask = df["topics"].apply(lambda t: topic in t)
        topic_total = int(topic_mask.sum())
        if topic_total == 0:
            continue
        topic_genuine_mask = topic_mask & (df["email_type"] == "genuine_question")
        topic_genuine = int(topic_genuine_mask.sum())
        pct_genuine = round(100 * topic_genuine / topic_total, 1) if topic_total else 0
        examples = df.loc[topic_genuine_mask, "new_content"].str.strip()
        examples = [e for e in examples if 15 < len(e) < 220][:3]
        rows.append(dict(
            topic=topic, total_emails=topic_total,
            pct_of_mailbox=round(100 * topic_total / total, 1),
            genuine_questions=topic_genuine, pct_genuine=pct_genuine,
            examples=examples,
        ))

    backlog = pd.DataFrame(rows, columns=["topic", "total_emails", "pct_of_mailbox", "genuine_questions", "pct_genuine", "examples"])
    backlog = backlog.sort_values("genuine_questions", ascending=False).reset_index(drop=True)
    if len(backlog):
        backlog.insert(0, "rank", range(1, len(backlog) + 1))

    type_summary = df["email_type"].value_counts().rename_axis("type").reset_index(name="count")
    type_summary["pct"] = round(100 * type_summary["count"] / total, 1)

    return df, backlog, type_summary


# ============================================================
# Semantic analysis (paid path)
# ============================================================
def run_llm_analysis(df, subject_col, body_col, api_key, model, sample_size=40,
                     batch_size=None, progress_callback=None, status_callback=None,
                     known_topics=None):
    """
    The AI-powered path: no pre-defined topic list needed -- discovers
    topics from the data itself, then classifies every email with real
    language understanding rather than keyword matching.

    known_topics: pass in topics already discovered during the scoping
    step, to avoid paying for and waiting on a second discovery call.

    Returns (df, backlog, type_summary, failed_batches). failed_batches is
    the number of classification batches that raised and were recorded as
    'unclear' -- the caller must disclose a non-zero count to the user, in
    the interface and in the downloadable report, because silently
    degraded runs otherwise present partial results as complete findings.
    """
    from llm_classifier import discover_topics_with_llm

    batch_size = batch_size or BATCH_SIZE
    client = make_client(api_key)

    df = df.copy()
    df[subject_col] = df[subject_col].fillna("").astype(str)
    df[body_col] = df[body_col].fillna("").astype(str).apply(strip_html).apply(strip_quotes)
    emails = [{"subject": r[subject_col], "body": r[body_col]} for _, r in df.iterrows()]

    if known_topics:
        topics = known_topics
    else:
        if status_callback:
            status_callback("Reading a sample to discover this mailbox's actual topics...")
        sample = emails[:min(sample_size, len(emails))]
        topics = discover_topics_with_llm(client, model, sample)

    results = []
    failed_batches = 0
    n = len(emails)
    for i in range(0, n, batch_size):
        batch = emails[i:i + batch_size]
        if status_callback:
            status_callback(f"Classifying emails {i+1}-{min(i+batch_size, n)} of {n}...")
        try:
            batch_results = classify_batch_with_llm(client, model, batch, topics)
        except Exception:
            # Don't let one bad batch crash the whole run -- mark it
            # unclear and keep going, but COUNT it so the degradation is
            # disclosed instead of silently absorbed into the findings.
            failed_batches += 1
            batch_results = [{"type": "unclear", "topic": "Other", "specificity": "not_applicable"}
                             for _ in batch]
        results.extend(batch_results)
        if progress_callback:
            progress_callback(min(1.0, (i + batch_size) / n))

    df["email_type"] = [r.get("type", "unclear") for r in results]
    df["topics"] = [[r.get("topic", "Other")] for r in results]
    df["specificity"] = [r.get("specificity", "not_applicable") for r in results]
    df["new_content"] = df[body_col]

    total = len(df)
    rows = []
    for topic in topics:
        topic_mask = df["topics"].apply(lambda t: topic in t)
        topic_total = int(topic_mask.sum())
        if topic_total == 0:
            continue
        genuine_mask = topic_mask & (df["email_type"] == "genuine_question")
        generic_mask = genuine_mask & (df["specificity"] == "generic")
        topic_generic = int(generic_mask.sum())
        # The ranking count AND the displayed percentage now share the same
        # numerator (article-addressable/generic questions). Previously the
        # count was generic-only while the percentage included case-specific
        # questions, so a card could read "5 questions ... 80%" with the two
        # figures measuring different things.
        pct_genuine = round(100 * topic_generic / topic_total, 1) if topic_total else 0
        examples = df.loc[generic_mask, body_col].str.strip()
        examples = [e[:220] for e in examples if 15 < len(e)][:3]
        rows.append(dict(
            topic=topic, total_emails=topic_total,
            pct_of_mailbox=round(100 * topic_total / total, 1),
            genuine_questions=topic_generic,  # ranked by article-addressable questions
            pct_genuine=pct_genuine, examples=examples,
        ))

    backlog = pd.DataFrame(rows, columns=["topic", "total_emails", "pct_of_mailbox", "genuine_questions", "pct_genuine", "examples"])
    backlog = backlog.sort_values("genuine_questions", ascending=False).reset_index(drop=True)
    if len(backlog):
        backlog.insert(0, "rank", range(1, len(backlog) + 1))

    type_summary = df["email_type"].value_counts().rename_axis("type").reset_index(name="count")
    type_summary["pct"] = round(100 * type_summary["count"] / total, 1)

    return df, backlog, type_summary, failed_batches


# ============================================================
# Presentation helpers
# ============================================================
def confidence_band(pct_genuine):
    if pct_genuine >= 70:
        return "High", "badge-high", "●"
    if pct_genuine >= 40:
        return "Medium", "badge-medium", "●"
    return "Needs review", "badge-low", "●"


def suggest_format(topic):
    heuristics = {
        "Password": "Step-by-step guide", "VPN": "Troubleshooting guide", "Software": "FAQ + request form",
        "Printer": "Troubleshooting guide", "New Starter": "Checklist", "Email": "FAQ",
        "WiFi": "FAQ", "Meeting": "Booking guide", "Security": "FAQ + reporting guide",
    }
    for key, fmt in heuristics.items():
        if key.lower() in topic.lower():
            return fmt
    return "FAQ article"


def sanitize_for_csv(df):
    """Neutralise spreadsheet formula injection in exported CSVs.

    Topic names and other string values in the backlog originate from the
    LLM (i.e. indirectly from untrusted mailbox content) or from user
    input. A value beginning with = + - or @ executes as a formula when
    the client opens the CSV in Excel. Prefix such strings with a
    single quote so they render as text."""
    out = df.copy()
    for col in out.columns:
        # pandas 2.x stores text as object dtype; pandas 3.x defaults to a
        # dedicated string dtype. Checking only `== object` silently skips
        # sanitisation on newer pandas, so test for both.
        if out[col].dtype == object or pd.api.types.is_string_dtype(out[col]):
            out[col] = out[col].map(
                lambda v: "'" + v if isinstance(v, str) and v[:1] in ("=", "+", "-", "@") else v)
    return out


# ============================================================
# Column auto-detection and mailbox validation
# ============================================================
def auto_detect_column(columns, must_contain_any):
    """Returns (best_guess, confident) -- confident=True means we found a
    strong match and don't need to bother the user about it."""
    cols_lower = {c: c.lower() for c in columns}
    # Pass 1: exact match on common names
    for target in must_contain_any:
        for c, cl in cols_lower.items():
            if cl == target:
                return c, True
    # Pass 2: contains match
    for target in must_contain_any:
        for c, cl in cols_lower.items():
            if target in cl:
                return c, True
    # No confident match found
    return None, False


def guess_body_by_content(df, exclude_cols):
    """When column names give no clues, the body column is reliably the
    one with the longest average text -- email bodies are almost always
    longer than subjects, senders, or dates."""
    candidates = [c for c in df.columns if c not in exclude_cols]
    if not candidates:
        return df.columns[0]
    avg_lengths = {c: df[c].astype(str).str.len().mean() for c in candidates}
    return max(avg_lengths, key=avg_lengths.get)


def guess_sender_by_content(df, exclude_cols):
    """When column names give no clues, a column where most values contain
    '@' is very likely a sender/email field."""
    candidates = [c for c in df.columns if c not in exclude_cols]
    for c in candidates:
        vals = df[c].astype(str)
        if (vals.str.contains("@")).mean() > 0.5:
            return c
    return None


def detect_all_columns(df):
    cols = df.columns.tolist()
    subject_col, subj_conf = auto_detect_column(cols, ["subject", "email subject"])
    body_col, body_conf = auto_detect_column(cols, ["body", "message", "content"])
    sender_col, sender_conf = auto_detect_column(cols, ["sender_name", "sender name", "from name", "sender", "from"])

    # Name-based detection can resolve subject and body to the SAME column
    # (e.g. a lone "Message Subject" column matches both target lists).
    # Treat the body match as unresolved in that case so the content-based
    # fallback assigns a different column.
    if body_col is not None and body_col == subject_col:
        body_col, body_conf = None, False

    # Seed the exclusion set with every confidently named column BEFORE the
    # content-based fallbacks run. Previously the fallback could select the
    # already-detected subject column as the body column, silently
    # double-counting the same text in every downstream metric.
    used = {c for c in (subject_col, body_col, sender_col) if c is not None}

    if body_col is None:
        body_col = guess_body_by_content(df, used)
    used.add(body_col)
    if sender_col is None:
        sender_col = guess_sender_by_content(df, used)
    if sender_col is not None:
        used.add(sender_col)
    if subject_col is None:
        remaining = [c for c in cols if c not in used]
        subject_col = remaining[0] if remaining else cols[0]

    # Sender is optional: a valid export may not contain it at all.
    all_confident = subj_conf and body_conf
    return dict(subject=subject_col, body=body_col, sender=sender_col), all_confident


def validate_mailbox_dataframe(df, detected, confident):
    """Rejects readable spreadsheets that do not contain usable email data.

    Named subject/body fields are strong evidence of a mailbox export --
    but not proof: a project tracker can also have a "Subject" column. A
    minimal message-likeness check therefore applies on BOTH paths, with a
    lower bar when column names matched confidently."""
    if df is None or df.empty or len(df.columns) == 0:
        return False, "The file does not contain any data rows."

    body_col = detected.get("body")
    subject_col = detected.get("subject")
    if body_col not in df.columns:
        return False, "A usable message-body field could not be identified."

    def clean_value(value):
        if pd.isna(value):
            return ""
        value = re.sub(r"\s+", " ", str(value)).strip()
        return "" if value.lower() in {"nan", "none", "null"} else value

    body_text = df[body_col].map(clean_value)
    subject_text = df[subject_col].map(clean_value) if subject_col in df.columns else pd.Series("", index=df.index)
    usable_mask = (body_text.str.len() > 0) | (subject_text.str.len() > 0)
    usable_count = int(usable_mask.sum())
    nonempty_body_count = int((body_text.str.len() > 0).sum())

    if usable_count == 0 or nonempty_body_count == 0:
        return False, "The file has no usable email subject or message content."

    message_like_count = int((body_text.str.len() >= 25).sum())
    substantial_count = int((body_text.str.len() >= 60).sum())

    if not confident:
        # Unknown column names are allowed, but the fallback body must resemble
        # a collection of messages rather than headings or spreadsheet labels.
        if message_like_count < 2 and substantial_count < 1:
            return False, (
                "The spreadsheet is readable, but it does not appear to contain mailbox data. "
                "No subject/body columns were recognised and there is not enough message-like text to analyze."
            )
    else:
        # Even with named columns, require minimal message-like evidence so a
        # tracker with a coincidental "Subject" column is not analysed as email.
        if message_like_count < 1 and nonempty_body_count < 3:
            return False, (
                "Subject and body columns were found by name, but the content does not resemble "
                "email messages. Check this is a mailbox export rather than a tracker or template."
            )

    return True, ""
